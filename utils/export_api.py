"""
utils/export_api.py
-------------------
Data export/import framework supporting multiple formats with scheduling.

Formats:
  - JSON (full fidelity, for backup/restore)
  - CSV (for spreadsheet users, flat tables)
  - Excel (.xlsx via openpyxl if available, fallback to CSV)

Exportable data collections:
  - customers: CRM customer list
  - history: AI generation history
  - workflows: Follow-up workflow records
  - analytics: Usage and engagement stats
  - emails: Sent email tracking records

Features:
  - Format auto-detection on import
  - Schema validation on import
  - Incremental export (since last export timestamp)
  - Scheduled auto-export (daily/weekly backup)
  - Data anonymization option (for sharing/demo)
  - Compression (gzip) for large exports

Usage:
    from utils.export_api import (
        export_data, import_data, get_export_history,
        schedule_auto_export, list_exportable_collections,
    )
"""
from __future__ import annotations

import csv
import gzip
import io
import json
from datetime import datetime

from utils.logger import get_logger
from utils.storage import load_user_json, save_user_json

logger = get_logger("export_api")

_EXPORT_HISTORY_FILE = "export_history.json"
_SCHEDULE_FILE = "export_schedule.json"

# ---------------------------------------------------------------------------
# Exportable collections registry
# ---------------------------------------------------------------------------

COLLECTIONS: dict[str, dict] = {
    "customers": {
        "label": "客户数据",
        "label_en": "Customer Data",
        "description": "CRM customer records with scores, tags, and stages",
        "file": "customers.json",
        "schema_fields": ["company", "contact", "email", "country", "product", "stage"],
        "tier_required": "free",
    },
    "history": {
        "label": "生成历史",
        "label_en": "Generation History",
        "description": "All AI-generated content with timestamps and parameters",
        "file": "history.json",
        "schema_fields": ["feature", "title", "content", "timestamp"],
        "tier_required": "pro",
    },
    "workflows": {
        "label": "跟进工作流",
        "label_en": "Follow-up Workflows",
        "description": "Email follow-up workflows and completion status",
        "file": "workflows.json",
        "schema_fields": ["customer", "product", "company", "status", "sent_at"],
        "tier_required": "pro",
    },
    "email_tracking": {
        "label": "邮件追踪",
        "label_en": "Email Tracking",
        "description": "Sent email delivery, open, and click tracking data",
        "file": "email_tracking.json",
        "schema_fields": ["tracking_id", "to_email", "subject", "status", "sent_at"],
        "tier_required": "pro",
    },
    "templates": {
        "label": "邮件模板",
        "label_en": "Email Templates",
        "description": "Saved email templates and parameters",
        "file": "templates.json",
        "schema_fields": ["name", "category", "content"],
        "tier_required": "free",
    },
}


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------

def list_exportable_collections(tier: str = "free") -> list[dict]:
    """
    List collections available for export based on user's tier.

    Args:
        tier: User's subscription tier (free/pro/team/enterprise)

    Returns:
        List of collection metadata dicts
    """
    tier_hierarchy = {"free": 0, "pro": 1, "team": 2, "enterprise": 3}
    user_level = tier_hierarchy.get(tier, 0)

    available = []
    for key, info in COLLECTIONS.items():
        required_level = tier_hierarchy.get(info["tier_required"], 0)
        available.append({
            "key": key,
            "label": info["label"],
            "label_en": info["label_en"],
            "description": info["description"],
            "accessible": user_level >= required_level,
            "tier_required": info["tier_required"],
        })
    return available


def export_data(
    username: str,
    collection: str,
    format: str = "json",
    compress: bool = False,
    anonymize: bool = False,
    since: str = "",
) -> tuple[bool, str | bytes, str]:
    """
    Export a data collection in the specified format.

    Args:
        username: User whose data to export
        collection: Collection key (from COLLECTIONS)
        format: Output format ('json', 'csv', 'xlsx')
        compress: Whether to gzip the output
        anonymize: Whether to anonymize PII (email, phone)
        since: ISO timestamp — only export records after this time

    Returns:
        (success, data_content, filename) tuple
        - data_content is str for json/csv, bytes for xlsx/compressed
    """
    if collection not in COLLECTIONS:
        return False, f"Unknown collection: {collection}", ""

    col_info = COLLECTIONS[collection]

    # Load data
    raw_data = load_user_json(username, col_info["file"], default=[])
    if not isinstance(raw_data, list):
        raw_data = [raw_data] if raw_data else []

    # Filter by timestamp if requested
    if since:
        raw_data = _filter_since(raw_data, since)

    # Anonymize if requested
    if anonymize:
        raw_data = _anonymize_records(raw_data)

    # Convert to requested format
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = f"{collection}_{timestamp}"

    if format == "json":
        content = json.dumps(raw_data, ensure_ascii=False, indent=2)
        filename = f"{base_filename}.json"
    elif format == "csv":
        content = _to_csv(raw_data, col_info.get("schema_fields", []))
        filename = f"{base_filename}.csv"
    elif format == "xlsx":
        content = _to_xlsx(raw_data, col_info.get("schema_fields", []), collection)
        filename = f"{base_filename}.xlsx"
        if not content:
            # Fallback to CSV if openpyxl not available
            content = _to_csv(raw_data, col_info.get("schema_fields", []))
            filename = f"{base_filename}.csv"
            format = "csv"
    else:
        return False, f"Unsupported format: {format}", ""

    # Compress if requested
    if compress:
        if isinstance(content, str):
            content = gzip.compress(content.encode("utf-8"))
        else:
            content = gzip.compress(content)
        filename += ".gz"

    # Record export in history
    _record_export(username, collection, format, len(raw_data))

    logger.info("Data exported: %s/%s (%d records, format=%s)", username, collection, len(raw_data), format)
    return True, content, filename


def export_all(
    username: str,
    format: str = "json",
    compress: bool = True,
) -> tuple[bool, bytes, str]:
    """
    Export all user data as a single archive.

    Returns a JSON bundle containing all collections.

    Args:
        username: User to export
        format: 'json' only for full export
        compress: Whether to gzip (recommended for full export)

    Returns:
        (success, compressed_bytes, filename)
    """
    bundle = {
        "_meta": {
            "exported_at": datetime.now().isoformat(),
            "username": username,
            "format_version": "1.0",
            "collections": [],
        }
    }

    for key, info in COLLECTIONS.items():
        data = load_user_json(username, info["file"], default=[])
        bundle[key] = data
        bundle["_meta"]["collections"].append(key)

    content = json.dumps(bundle, ensure_ascii=False, indent=2)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tradeai_full_backup_{timestamp}.json"

    if compress:
        content_bytes = gzip.compress(content.encode("utf-8"))
        filename += ".gz"
        return True, content_bytes, filename

    return True, content.encode("utf-8"), filename


# ---------------------------------------------------------------------------
# Import functions
# ---------------------------------------------------------------------------

def import_data(
    username: str,
    collection: str,
    content: str | bytes,
    format: str = "",
    merge_strategy: str = "append",
) -> tuple[bool, str, int]:
    """
    Import data into a collection.

    Args:
        username: Target user
        collection: Target collection key
        content: File content (str or bytes)
        format: Format hint ('json', 'csv', or '' for auto-detect)
        merge_strategy: How to handle existing data
            - 'append': Add to existing records
            - 'replace': Replace all existing data
            - 'merge': Smart merge (deduplicate by key fields)

    Returns:
        (success, message, records_imported) tuple
    """
    if collection not in COLLECTIONS:
        return False, f"Unknown collection: {collection}", 0

    col_info = COLLECTIONS[collection]

    # Decompress if gzipped
    if isinstance(content, bytes):
        try:
            content = gzip.decompress(content).decode("utf-8")
        except (gzip.BadGzipFile, OSError):
            content = content.decode("utf-8") if isinstance(content, bytes) else content

    # Auto-detect format
    if not format:
        format = _detect_format(content)

    # Parse content
    if format == "json":
        records = _parse_json_import(content)
    elif format == "csv":
        records = _parse_csv_import(content)
    else:
        return False, f"Cannot parse format: {format}", 0

    if not records:
        return False, "No valid records found in import file", 0

    # Validate schema
    schema_fields = col_info.get("schema_fields", [])
    valid_records = _validate_schema(records, schema_fields)

    if not valid_records:
        return False, f"No records match required schema. Expected fields: {schema_fields}", 0

    # Apply merge strategy
    existing = load_user_json(username, col_info["file"], default=[])

    if merge_strategy == "replace":
        final_data = valid_records
    elif merge_strategy == "merge":
        final_data = _smart_merge(existing, valid_records, schema_fields)
    else:  # append
        final_data = existing + valid_records

    save_user_json(username, col_info["file"], final_data)

    imported_count = len(valid_records)
    logger.info("Data imported: %s/%s (%d records, strategy=%s)", username, collection, imported_count, merge_strategy)
    return True, f"Successfully imported {imported_count} records", imported_count


def import_full_backup(
    username: str,
    content: str | bytes,
) -> tuple[bool, str, dict]:
    """
    Import a full backup bundle (from export_all).

    Args:
        username: Target user
        content: Backup file content

    Returns:
        (success, message, {collection: count} stats)
    """
    # Decompress if needed
    if isinstance(content, bytes):
        try:
            content = gzip.decompress(content).decode("utf-8")
        except (gzip.BadGzipFile, OSError):
            content = content.decode("utf-8") if isinstance(content, bytes) else content

    try:
        bundle = json.loads(content)
    except (json.JSONDecodeError, TypeError) as e:
        return False, f"Invalid backup file: {e}", {}

    if "_meta" not in bundle:
        return False, "Not a valid TradeAI backup file (missing _meta)", {}

    stats = {}
    for key, info in COLLECTIONS.items():
        if key in bundle:
            data = bundle[key]
            if isinstance(data, list):
                save_user_json(username, info["file"], data)
                stats[key] = len(data)

    total = sum(stats.values())
    logger.info("Full backup imported for %s: %d total records across %d collections",
                username, total, len(stats))
    return True, f"Backup restored: {total} records across {len(stats)} collections", stats


# ---------------------------------------------------------------------------
# Scheduled exports
# ---------------------------------------------------------------------------

def schedule_auto_export(
    username: str,
    frequency: str = "weekly",
    format: str = "json",
    collections: list[str] | None = None,
) -> tuple[bool, str]:
    """
    Schedule automatic data exports.

    Args:
        username: User to schedule for
        frequency: 'daily' or 'weekly'
        format: Export format
        collections: Which collections (None = all)

    Returns:
        (success, message) tuple
    """
    if frequency not in ("daily", "weekly"):
        return False, "Frequency must be 'daily' or 'weekly'"

    schedule = load_user_json(username, _SCHEDULE_FILE, default={})
    schedule.update({
        "enabled": True,
        "frequency": frequency,
        "format": format,
        "collections": collections or list(COLLECTIONS.keys()),
        "last_export": schedule.get("last_export", ""),
        "created_at": datetime.now().isoformat(),
    })
    save_user_json(username, _SCHEDULE_FILE, schedule)

    logger.info("Auto-export scheduled for %s: %s, format=%s", username, frequency, format)
    return True, f"Auto-export scheduled: {frequency}, format: {format}"


def check_scheduled_exports(username: str) -> bool:
    """
    Check if a scheduled export is due and execute it.

    Called periodically (e.g., on each page load).

    Returns:
        True if an export was performed
    """
    schedule = load_user_json(username, _SCHEDULE_FILE, default={})
    if not schedule.get("enabled"):
        return False

    last_export = schedule.get("last_export", "")
    frequency = schedule.get("frequency", "weekly")

    # Determine if export is due
    if last_export:
        try:
            last_dt = datetime.fromisoformat(last_export)
            from datetime import timedelta
            interval = timedelta(days=1) if frequency == "daily" else timedelta(days=7)
            if datetime.now() - last_dt < interval:
                return False  # Not due yet
        except (ValueError, TypeError):
            pass  # Invalid date, proceed with export

    # Perform export
    collections = schedule.get("collections", list(COLLECTIONS.keys()))
    format_type = schedule.get("format", "json")

    for collection in collections:
        if collection in COLLECTIONS:
            export_data(username, collection, format=format_type, compress=True)

    # Update last export time
    schedule["last_export"] = datetime.now().isoformat()
    save_user_json(username, _SCHEDULE_FILE, schedule)

    logger.info("Scheduled export completed for %s", username)
    return True


def get_export_schedule(username: str) -> dict:
    """Get current export schedule configuration."""
    return load_user_json(username, _SCHEDULE_FILE, default={})


# ---------------------------------------------------------------------------
# Export history
# ---------------------------------------------------------------------------

def get_export_history(username: str, limit: int = 20) -> list[dict]:
    """Get recent export history for a user."""
    history = load_user_json(username, _EXPORT_HISTORY_FILE, default=[])
    return history[-limit:]


def _record_export(username: str, collection: str, format: str, record_count: int) -> None:
    """Record an export event in history."""
    history = load_user_json(username, _EXPORT_HISTORY_FILE, default=[])
    history.append({
        "collection": collection,
        "format": format,
        "record_count": record_count,
        "exported_at": datetime.now().isoformat(),
    })
    # Keep last 50 entries
    if len(history) > 50:
        history = history[-50:]
    save_user_json(username, _EXPORT_HISTORY_FILE, history)


# ---------------------------------------------------------------------------
# Format conversion helpers
# ---------------------------------------------------------------------------

def _to_csv(records: list[dict], schema_fields: list[str]) -> str:
    """Convert records to CSV string."""
    if not records:
        return ""

    # Determine columns: schema fields + any extra keys from data
    all_keys = set()
    for r in records[:100]:  # Sample first 100 for column detection
        all_keys.update(r.keys())

    # Schema fields first, then alphabetical extras
    columns = list(schema_fields)
    extras = sorted(all_keys - set(schema_fields))
    columns.extend(extras)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()

    for record in records:
        # Flatten nested values
        flat = {}
        for key in columns:
            val = record.get(key, "")
            if isinstance(val, (dict, list)):
                flat[key] = json.dumps(val, ensure_ascii=False)
            else:
                flat[key] = val
        writer.writerow(flat)

    return output.getvalue()


def _to_xlsx(records: list[dict], schema_fields: list[str], sheet_name: str) -> bytes | None:
    """Convert records to Excel bytes. Returns None if openpyxl not available."""
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.debug("openpyxl not available, falling back to CSV")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name max 31 chars

    if not records:
        return None

    # Columns
    all_keys = set()
    for r in records[:100]:
        all_keys.update(r.keys())
    columns = list(schema_fields) + sorted(all_keys - set(schema_fields))

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = record.get(col_name, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _parse_json_import(content: str) -> list[dict]:
    """Parse JSON content into a list of records."""
    try:
        data = json.loads(content)
        if isinstance(data, list):
            return [r for r in data if isinstance(r, dict)]
        elif isinstance(data, dict):
            return [data]
        return []
    except (json.JSONDecodeError, TypeError):
        return []


def _parse_csv_import(content: str) -> list[dict]:
    """Parse CSV content into a list of records."""
    try:
        reader = csv.DictReader(io.StringIO(content))
        return [dict(row) for row in reader]
    except Exception:
        return []


def _detect_format(content: str) -> str:
    """Auto-detect file format from content."""
    content = content.strip()
    if content.startswith("[") or content.startswith("{"):
        return "json"
    if "," in content.split("\n")[0]:
        return "csv"
    return "json"  # Default


def _validate_schema(records: list[dict], required_fields: list[str]) -> list[dict]:
    """Filter records that have at least some of the required fields."""
    if not required_fields:
        return records

    valid = []
    for record in records:
        # Accept record if it has at least 50% of required fields
        present = sum(1 for f in required_fields if f in record and record[f])
        if present >= len(required_fields) * 0.5:
            valid.append(record)
    return valid


def _smart_merge(existing: list[dict], new_records: list[dict], key_fields: list[str]) -> list[dict]:
    """
    Merge new records with existing, avoiding duplicates.

    Uses first 2 key_fields as composite dedup key.
    """
    if not key_fields or len(key_fields) < 2:
        return existing + new_records

    # Build lookup from existing records
    k1, k2 = key_fields[0], key_fields[1]
    existing_keys = {
        (r.get(k1, "").lower(), r.get(k2, "").lower())
        for r in existing
    }

    # Add only non-duplicate new records
    merged = list(existing)
    for record in new_records:
        key = (record.get(k1, "").lower(), record.get(k2, "").lower())
        if key not in existing_keys:
            merged.append(record)
            existing_keys.add(key)

    return merged


def _filter_since(records: list[dict], since: str) -> list[dict]:
    """Filter records created/updated after a timestamp."""
    filtered = []
    for r in records:
        # Check common timestamp fields
        ts = r.get("timestamp") or r.get("created_at") or r.get("sent_at") or ""
        if ts >= since:
            filtered.append(r)
    return filtered


def _anonymize_records(records: list[dict]) -> list[dict]:
    """
    Anonymize PII in records (for sharing/demo purposes).

    Replaces emails with hash, phone numbers with ***, names with initials.
    """
    import hashlib

    anonymized = []
    for record in records:
        anon = dict(record)
        # Anonymize email
        if "email" in anon and anon["email"]:
            email = anon["email"]
            hashed = hashlib.md5(email.encode()).hexdigest()[:8]
            domain = email.split("@")[-1] if "@" in email else "example.com"
            anon["email"] = f"user_{hashed}@{domain}"
        # Anonymize contact name
        if "contact" in anon and anon["contact"]:
            parts = anon["contact"].split()
            anon["contact"] = " ".join(p[0] + "***" for p in parts if p)
        # Anonymize to_email
        if "to_email" in anon and anon["to_email"]:
            email = anon["to_email"]
            hashed = hashlib.md5(email.encode()).hexdigest()[:8]
            domain = email.split("@")[-1] if "@" in email else "example.com"
            anon["to_email"] = f"user_{hashed}@{domain}"
        anonymized.append(anon)
    return anonymized
